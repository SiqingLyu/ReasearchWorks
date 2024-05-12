# coding=UTF-8
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy
from openpyxl import Workbook


def replace_ascs(s):
    ascs=['\xc2', '\xa0']
    for asc in ascs:
        s = s.replace(asc, '')
    return s


def read_xls_as_list(path, sheet):
    data = read_excel_xls(path, sheet)
    rowAmount = data.nrows
    colAmount = data.ncols
    info_list = [[]]
    for rowIndex in range(rowAmount):
        for colIndex in range(colAmount):
            s = str(data.cell_value(rowIndex, colIndex))
            info_list[rowIndex].append(replace_ascs(s))
            # info_list[rowIndex].append(s)
            if ('\\' in s) & ('\n' not in s):
                print(s + "has unknown chars")
        info_list.append([])
    return info_list


def write_excel_xls(path, sheet_name, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
    workbook.save(path)  # 保存工作簿
    # print("xls格式表格写入数据成功！")


def write_excel_xls_append(path, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            new_worksheet.write(i + rows_old, j, value[i][j])  # 追加写入数据，注意是从i+rows_old行开始写入
    new_workbook.save(path)  # 保存工作簿
    # print("xls格式表格【追加】写入数据成功！")


def read_excel_xls(path, sheet_name):
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheet_name)  # 获取工作簿中所有表格中的的第一个表格
    return worksheet
    # for i in range(0, worksheet.nrows):
    #     for j in range(0, worksheet.ncols):
    #         print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
    #     print()


def write_excel_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    workbook.save(path)
    # print("xlsx格式表格写入数据成功！")


def read_excel_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    # sheet = wb.get_sheet_by_name(sheet_name)这种方式已经弃用，不建议使用
    sheet = workbook[sheet_name]
    return sheet
    # for row in sheet.rows:
    #     for cell in row:
    #         print(cell.value, "\t", end="")
    #     print()


def add_sheet(path, value, sheet):
    '''
	:param sheet:sheet的名称
    :param path:写入excel的路径
    :param value: 追加的数据
    :return:
    '''
    wb = openpyxl.load_workbook(path)
    wb.create_sheet(sheet)
    ws = wb[sheet]

    for ss in value:
        ws.append(ss)
    wb.save(path)
    # print("写入成功")


